import pandas
from utils.case_conf.case_conf import GetConf
from settings import CONF_INI
from settings import EXCEL_FILE
from openpyxl import load_workbook
from utils.do_mysql import DoMysql
import random

class DoExcel:

    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)
        self.init_tel = None

    def read_value(self):
        self.df = pandas.read_excel(self.file_name,'init')
        print(self.df.ix[1,1])

    def get_conf(self):
        conf_value = GetConf(CONF_INI,'CASE_MODE','case_mode').get_conf()
        return conf_value

    def get_loanid(self):
        loan_id = DoMysql().execute_sql('select id from loan where MemberId = 11')
        return loan_id

    def replace_data(self,case_info,sheet_name):
        if case_info['case_data'].find('${tel}') != -1:
            self.init_tel = int(pandas.read_excel(self.file_name, 'init').ix[0, 1])
            print(self.init_tel)
            case_info['case_data'] = case_info['case_data'].replace('${tel}', str(self.init_tel))
        if case_info['case_data'].find('${admin_tel}') != -1:
            self.admin_tel = int(pandas.read_excel(self.file_name, 'init').ix[1, 1])
            case_info['case_data'] = case_info['case_data'].replace('${admin_tel}', str(self.admin_tel))
        if case_info['case_data'].find('${invest_memid}') != -1:
            self.invest_memid = int(pandas.read_excel(self.file_name, 'init').ix[3, 1])
            case_info['case_data'] = case_info['case_data'].replace('${invest_memid}', str(self.invest_memid))
        if case_info['case_data'].find('${loan_id}')!=-1:
            loan_id = self.get_loanid()
            case_info['case_data'] = case_info['case_data'].replace('${loan_id}', str(loan_id))
        case_info['sheet_name'] = sheet_name

    def get_case_list(self):
        case_list = []
        mode_info = self.get_conf()
        for sheet_name,case_mode in eval(mode_info).items():
            # index_col表示索引列，0为第一列
            self.df = pandas.read_excel(self.file_name,sheet_name)
            if case_mode == 'all':
                for row in self.df.index.values:
                    case_info = self.df.ix[row].to_dict()
                    self.replace_data(case_info,sheet_name)
                    case_list.append(case_info)
            elif case_mode == 'random':
                rand_int = random.choice([i for i in range(1,max(self.df.index.values)+2)])
                for case in self.df.sample(rand_int).values:
                    case_info = self.df.ix[case[0]-1].to_dict()
                    self.replace_data(case_info,sheet_name)
                    case_list.append(case_info)
            else:
                for row in case_mode:
                    case_info = self.df.ix[row-1].to_dict()
                    self.replace_data(case_info,sheet_name)
                    case_list.append(case_info)
        # 最后保存并关闭文件
        if self.init_tel:
            self.update_tel('init',self.init_tel)
        print(case_list)
        return case_list

    def update_tel(self,sheet_name,new_tel):
        sheet = self.wb[sheet_name]
        # sheet.cell(2,2).value = new_tel+1
        # self.wb.save(self.file_name)

    def update_success(self,sheet_name,is_sucess,case_info):
        sheet = self.wb[sheet_name]
        sheet.cell(case_info['case_id'] + 1, 8).value = is_sucess
        self.wb.save(self.file_name)

    def update_result(self,sheet_name,result,case_info):
        sheet = self.wb[sheet_name]
        sheet.cell(case_info['case_id'] + 1, 7).value = str(result)
        self.wb.save(self.file_name)

    def update_amount(self,sheet_name,amount_result,case_info):
        sheet = self.wb[sheet_name]
        sheet.cell(case_info['case_id'] + 1, 9).value = str(amount_result)
        self.wb.save(self.file_name)


if __name__ == '__main__':
    DoExcel(EXCEL_FILE).get_case_list()


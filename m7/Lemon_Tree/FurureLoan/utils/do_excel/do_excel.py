from openpyxl import load_workbook
from utils.case_conf.case_conf import GetConf
from settings import CONF_INI

class DoExcel:
    api_url = 'http://47.107.168.87:8080/futureloan/mvc/api/member/'

    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)

    def get_mode(self):
        case_mode = GetConf(CONF_INI, 'CASE_MODE', 'case_mode').get_conf()
        return eval(case_mode)

    def set_case_info(self,sheet,row,mode,case_list):
        case_info = {}
        case_info['case_id'] = sheet.cell(row, 1).value
        case_info['case_title'] = sheet.cell(row, 2).value
        case_info['case_data'] = sheet.cell(row, 3).value
        case_info['case_expected'] = sheet.cell(row, 4).value
        case_info['url'] = self.api_url + mode
        case_info['sheet_name'] = mode
        case_list.append(case_info)

    def get_case_list(self):
        case_mode = self.get_mode()
        case_list = []
        for mode in case_mode:
            sheet = self.wb[mode]
            if case_mode[mode] == 'all':
                for row in range(2,sheet.max_row+1):
                    self.set_case_info(sheet,row,mode,case_list)
            else:
                for row in case_mode[mode]:
                    row=int(row)+1
                    self.set_case_info(sheet,row,mode,case_list)
        return case_list

    def set_value(self,sheet_name,row,result,is_sucess):
        sheet = self.wb[sheet_name]
        sheet.cell(row+1,5).value = result
        sheet.cell(row+1,6).value = is_sucess
        self.wb.save(self.file_name)

if __name__ == '__main__':
    DoExcel('test_case.xlsx').get_case_list()
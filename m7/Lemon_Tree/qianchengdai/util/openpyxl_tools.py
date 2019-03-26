from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name,button = 'all'):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.button = button
        self.sheet = load_workbook(self.file_name)[self.sheet_name]

    def get_case_list(self):
        case_list = []
        final_case_list = []

        for row in range(2,self.sheet.max_row+1):
            case_data = {}
            case_data['case_id'] = self.sheet.cell(row,1).value
            case_data['case_desc'] = self.sheet.cell(row,2).value
            case_data['case_data'] = self.sheet.cell(row,4).value
            case_data['case_expected'] = self.sheet.cell(row,5).value
            case_list.append(case_data)

        if self.button == 'all':
            final_case_list = case_list
        else:
            for id in eval(self.button):
                for case in case_list:
                    if case['case_id'] == id:
                        final_case_list.append(case)

        return final_case_list